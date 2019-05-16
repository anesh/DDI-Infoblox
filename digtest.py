import openpyxl
import xlsxwriter
import requests
import json
requests.packages.urllib3.disable_warnings()


book = xlsxwriter.Workbook('p0adqppdhcp00.ns.ctc1.xlsx')
sheet1 = book.add_worksheet("report")

row=0
col=0


workbook = openpyxl.load_workbook('delegationsv2.xlsx')
sheet = workbook.active
row_count = sheet.max_row
col_count = sheet.max_column
try:
	for i in range(2,row_count+1):
		cellobjrow=sheet.cell(row=i,column=1)
		fqdn = cellobjrow.value
		print "FQDN is:"+fqdn
		for j in range(2,col_count+1):
			row=row+1
			col=col+1
			cellobj=sheet.cell(row=i,column=j)
			nsrecord = cellobj.value
			sheet1.write(row, col,fqdn)
			col=col+1
			sheet1.write(row, col,nsrecord) 
			data={"fqdn" : str(fqdn) , "member": "p0adqppdhcp00.ns.ctc", "name_server": str(nsrecord).rstrip("."), "record_type":"ANY"}
			url= "https://10.2.61.100/wapi/v2.7/grid/b25lLmNsdXN0ZXIkMA:Infoblox?_function=query_fqdn_on_member"
			payload = json.dumps(data) 
			headers = {'content-type': "application/json"}
			response = requests.request("POST", url, auth=('anesh.ponnarasseryke', 'Adimurai@777'), data=payload, headers=headers,verify=False)
			print response.text
			if response.status_code == 200:
				col=col+1
				sheet1.write(row, col,response.text)
                	col=0

			if cellobj.value is None:
				break
except Exception as e:
	print e

book.close()
